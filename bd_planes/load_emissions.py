#Модуль посвящен созданию и загрузки выбросов самолетов в БД из файла
import psycopg2
from openpyxl import load_workbook


def get_data_from_file():
    print(1)
    wb = load_workbook('emission_data.xlsx', data_only=True)
    ws = wb.active

    data_arr = []

    for cellO in ws['B3':'T147']:
        data_struct = {}
        for cell in cellO:
            k = str(cell)
            d = k.rfind('.')
            k = k[d:]
            if 'B' in k:
                data_struct['id'] = cell.value
            if 'C' in k:
                data_struct['Plane'] = cell.value
            if 'I' in k:
                passen = cell.value
            if 'J' in k:
                data_struct['distance'] = cell.value
            if 'N' in k:
                data_struct['fuel_value'] = cell.value
            if 'P' in k:
                data_struct['emission'] = cell.value
            if 'R' in k:
                data_struct['distance_per_hundred'] = cell.value
                data_struct['emission_per_passenger'] = cell.value / passen
            if 'T' in k:
                data_struct['useful_weight'] = cell.value
        data_arr.append(data_struct)

    #print(len(data_arr))
    #print(data_arr[137])
    return data_arr


def load_data_to_bd(data_arr):
    conn = psycopg2.connect(database="BD_Planes", user='postgres',
                            password='YUV45G', host='localhost', port=5432)
    cur = conn.cursor()
    try:
        cur.execute("CREATE TABLE Emissions"
                    "("
                    "Plane varchar(30) primary key, "
                    "distance Numeric(7) not null, "
                    "fuel_value Numeric(10) not null, "
                    "useful_weight Numeric(10) not null, "
                    "emission Numeric(10) not null, "
                    "emission_per_hundred_km Numeric(10) not null, "
                    "emission_per_passenger_per_100km Numeric(6,4) not null "
                    ")")
        conn.commit()
    except psycopg2.errors.DuplicateTable:
        cur.execute("ROLLBACK")

    for data in data_arr:
        try:
            cur.execute("INSERT INTO Emissions VALUES (%s, %s, %s, %s, %s, %s, %s)", (data['Plane'], str(data['distance']), str(data['fuel_value']), str(data['useful_weight']), str(data['emission']), str(data['distance_per_hundred']), str(data['emission_per_passenger'])))
            conn.commit()
        except psycopg2.errors.UniqueViolation:
            cur.execute("ROLLBACK")

data_arr = get_data_from_file()
load_data_to_bd(data_arr)
